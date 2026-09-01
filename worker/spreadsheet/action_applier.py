"""Closed spreadsheet action application for the isolated LibreOffice worker."""

from __future__ import annotations

import copy
from collections.abc import Mapping
from typing import Any

from ananta_contracts.spreadsheet_studio import canonical_digest, cell_coordinates


class SpreadsheetActionApplier:
    """Apply already validated actions without owning execution or orchestration."""

    def apply(
        self,
        *,
        workbook: Any,
        sheet_names: Mapping[str, str],
        actions: tuple[Mapping[str, Any], ...],
        formula_asts: dict[tuple[str, str], Mapping[str, Any]],
        style_refs: dict[tuple[str, str], str | None],
    ) -> dict[tuple[str, str], list[str]]:
        direct: dict[tuple[str, str], list[str]] = {}
        for action in actions:
            kind = str(action["kind"])
            action_id = str(action["action_id"])
            if kind in {"set_value", "set_formula", "clear_cell"}:
                self._single_cell(
                    workbook=workbook,
                    sheet_names=sheet_names,
                    action=action,
                    formula_asts=formula_asts,
                    style_refs=style_refs,
                )
                _mark(direct, (str(action["sheet_id"]), str(action["cell"])), action_id)
            elif kind == "clear_range":
                sheet_id = str(action["sheet_id"])
                for address in _range_addresses(str(action["start"]), str(action["end"])):
                    workbook[sheet_names[sheet_id]][address].value = None
                    formula_asts.pop((sheet_id, address), None)
                    _mark(direct, (sheet_id, address), action_id)
            elif kind == "copy_range":
                self._copy_range(
                    workbook=workbook,
                    sheet_names=sheet_names,
                    action=action,
                    formula_asts=formula_asts,
                    style_refs=style_refs,
                    direct=direct,
                )
            elif kind == "format_range":
                self._format_range(
                    workbook=workbook,
                    sheet_names=sheet_names,
                    action=action,
                    style_refs=style_refs,
                    direct=direct,
                )
            else:
                self._structure(
                    workbook=workbook,
                    sheet_names=sheet_names,
                    action=action,
                    formula_asts=formula_asts,
                    style_refs=style_refs,
                    direct=direct,
                )
        return direct

    @staticmethod
    def _single_cell(
        *,
        workbook: Any,
        sheet_names: Mapping[str, str],
        action: Mapping[str, Any],
        formula_asts: dict[tuple[str, str], Mapping[str, Any]],
        style_refs: dict[tuple[str, str], str | None],
    ) -> None:
        from worker.spreadsheet.formula_parser import render_formula

        sheet_id = str(action["sheet_id"])
        address = str(action["cell"])
        target = workbook[sheet_names[sheet_id]][address]
        if action["kind"] == "clear_cell":
            target.value = None
            formula_asts.pop((sheet_id, address), None)
        elif action["kind"] == "set_formula":
            formula_asts[(sheet_id, address)] = copy.deepcopy(action["formula"])
            target.value = "=" + render_formula(action["formula"], sheet_names)
        else:
            target.value = action["value"]
            formula_asts.pop((sheet_id, address), None)
        style_refs.setdefault((sheet_id, address), None)

    @staticmethod
    def _copy_range(
        *,
        workbook: Any,
        sheet_names: Mapping[str, str],
        action: Mapping[str, Any],
        formula_asts: dict[tuple[str, str], Mapping[str, Any]],
        style_refs: dict[tuple[str, str], str | None],
        direct: dict[tuple[str, str], list[str]],
    ) -> None:
        source_sheet_id = str(action["source_sheet_id"])
        target_sheet_id = str(action["target_sheet_id"])
        source_sheet = workbook[sheet_names[source_sheet_id]]
        target_sheet = workbook[sheet_names[target_sheet_id]]
        source_start_row, source_start_column = cell_coordinates(action["source_start"])
        source_end_row, source_end_column = cell_coordinates(action["source_end"])
        target_start_row, target_start_column = cell_coordinates(action["target_start"])
        values = []
        for row_offset, source_row in enumerate(range(source_start_row, source_end_row + 1)):
            row = []
            for column_offset, source_column in enumerate(range(source_start_column, source_end_column + 1)):
                source = source_sheet.cell(source_row, source_column)
                source_address = source.coordinate
                row.append(
                    (
                        copy.copy(source.value),
                        copy.copy(source._style),  # noqa: SLF001 - openpyxl has no public style-copy API
                        copy.deepcopy(formula_asts.get((source_sheet_id, source_address))),
                        style_refs.get((source_sheet_id, source_address)),
                        row_offset,
                        column_offset,
                    )
                )
            values.append(row)
        for row in values:
            for value, style, formula, style_ref, row_offset, column_offset in row:
                target = target_sheet.cell(target_start_row + row_offset, target_start_column + column_offset)
                target.value = value
                target._style = style  # noqa: SLF001 - exact immutable style-array copy
                key = (target_sheet_id, target.coordinate)
                if formula is None:
                    formula_asts.pop(key, None)
                else:
                    formula_asts[key] = formula
                style_refs[key] = style_ref
                _mark(direct, key, str(action["action_id"]))

    @staticmethod
    def _format_range(
        *,
        workbook: Any,
        sheet_names: Mapping[str, str],
        action: Mapping[str, Any],
        style_refs: dict[tuple[str, str], str | None],
        direct: dict[tuple[str, str], list[str]],
    ) -> None:
        from openpyxl.styles import Font, PatternFill

        sheet_id = str(action["sheet_id"])
        sheet = workbook[sheet_names[sheet_id]]
        style = dict(action["style"])
        style_ref = f"style-{canonical_digest(style)[:24]}"
        for address in _range_addresses(str(action["start"]), str(action["end"])):
            cell = sheet[address]
            if style["number_format"] is not None:
                cell.number_format = str(style["number_format"])
            if style["bold"] is not None or style["italic"] is not None:
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.sz,
                    bold=cell.font.bold if style["bold"] is None else bool(style["bold"]),
                    italic=cell.font.italic if style["italic"] is None else bool(style["italic"]),
                    color=copy.copy(cell.font.color),
                )
            if style["fill_color"] is not None:
                cell.fill = PatternFill(fill_type="solid", fgColor=str(style["fill_color"]))
            style_refs[(sheet_id, address)] = style_ref
            _mark(direct, (sheet_id, address), str(action["action_id"]))

    @staticmethod
    def _structure(
        *,
        workbook: Any,
        sheet_names: Mapping[str, str],
        action: Mapping[str, Any],
        formula_asts: dict[tuple[str, str], Mapping[str, Any]],
        style_refs: dict[tuple[str, str], str | None],
        direct: dict[tuple[str, str], list[str]],
    ) -> None:
        sheet_id = str(action["sheet_id"])
        sheet = workbook[sheet_names[sheet_id]]
        kind = str(action["kind"])
        count = int(action["count"])
        if kind in {"insert_rows", "delete_rows"}:
            start = int(action["start_row"])
            if kind == "insert_rows":
                sheet.insert_rows(start, count)
            else:
                sheet.delete_rows(start, count)
            _rebase_metadata(
                values=formula_asts,
                style_refs=style_refs,
                sheet_id=sheet_id,
                axis="row",
                start=start,
                count=count,
                deleting=kind == "delete_rows",
            )
        else:
            start = int(action["start_column"])
            if kind == "insert_columns":
                sheet.insert_cols(start, count)
            else:
                sheet.delete_cols(start, count)
            _rebase_metadata(
                values=formula_asts,
                style_refs=style_refs,
                sheet_id=sheet_id,
                axis="column",
                start=start,
                count=count,
                deleting=kind == "delete_columns",
            )
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    _mark(direct, (sheet_id, cell.coordinate), str(action["action_id"]))


def _range_addresses(start: str, end: str) -> tuple[str, ...]:
    from openpyxl.utils import get_column_letter

    start_row, start_column = cell_coordinates(start)
    end_row, end_column = cell_coordinates(end)
    return tuple(
        f"{get_column_letter(column)}{row}"
        for row in range(start_row, end_row + 1)
        for column in range(start_column, end_column + 1)
    )


def _mark(direct: dict[tuple[str, str], list[str]], key: tuple[str, str], action_id: str) -> None:
    direct.setdefault(key, []).append(action_id)


def _rebase_metadata(
    *,
    values: dict[tuple[str, str], Mapping[str, Any]],
    style_refs: dict[tuple[str, str], str | None],
    sheet_id: str,
    axis: str,
    start: int,
    count: int,
    deleting: bool,
) -> None:
    from openpyxl.utils import get_column_letter

    def shifted(address: str) -> str | None:
        row, column = cell_coordinates(address)
        coordinate = row if axis == "row" else column
        if deleting and start <= coordinate < start + count:
            return None
        if coordinate >= start + (count if deleting else 0):
            coordinate += -count if deleting else count
        if axis == "row":
            row = coordinate
        else:
            column = coordinate
        return f"{get_column_letter(column)}{row}"

    for mapping in (values, style_refs):
        rebased = {}
        for (current_sheet, address), value in mapping.items():
            destination = shifted(address) if current_sheet == sheet_id else address
            if destination is not None:
                rebased[(current_sheet, destination)] = value
        mapping.clear()
        mapping.update(rebased)


__all__ = ["SpreadsheetActionApplier"]
