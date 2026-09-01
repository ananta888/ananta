"""Bounded LibreOffice-backed implementation of the spreadsheet execution port."""

from __future__ import annotations

import base64
import copy
import datetime as dt
import hashlib
import resource
import shutil
import subprocess
import tempfile
from collections.abc import Mapping
from pathlib import Path
from typing import Any

from ananta_contracts.spreadsheet_studio import WorkbookSnapshotV1, canonical_digest
from worker.spreadsheet.action_applier import SpreadsheetActionApplier
from worker.spreadsheet.artifact_inspector import SpreadsheetArtifactInspector
from worker.spreadsheet.formula_parser import SpreadsheetFormulaUnsupported, parse_formula, render_formula


class SpreadsheetExecutionError(RuntimeError):
    pass


class LibreOfficeSpreadsheetExecutor:
    """Runs one transformation in a fresh profile; orchestration remains in the Hub."""

    def __init__(
        self,
        *,
        executable: str | None = None,
        timeout_seconds: int = 90,
        memory_bytes: int = 2 * 1024**3,
        file_bytes: int = 256 * 1024**2,
        network_isolated: bool = False,
    ) -> None:
        resolved = executable or shutil.which("libreoffice") or shutil.which("soffice")
        if not resolved or not Path(resolved).is_absolute():
            raise SpreadsheetExecutionError("spreadsheet_libreoffice_unavailable")
        if not 1 <= int(timeout_seconds) <= 300:
            raise ValueError("spreadsheet_executor_timeout_invalid")
        self._executable = resolved
        self._timeout = int(timeout_seconds)
        self._memory_bytes = int(memory_bytes)
        self._file_bytes = int(file_bytes)
        self._network_isolated = bool(network_isolated)
        self._inspector = SpreadsheetArtifactInspector(max_compressed_bytes=min(self._file_bytes, 16 * 1024**2))
        self._actions = SpreadsheetActionApplier()
        self._version = self._detect_version()

    @property
    def capability(self) -> Mapping[str, Any]:
        return {
            "schema": "ananta.spreadsheet-executor-capability.v1",
            "state": "available",
            "engine": "libreoffice-calc",
            "engine_version": self._version,
            "network_enabled": not self._network_isolated,
            "macros_enabled": False,
            "external_links_enabled": False,
            "production_fidelity": self._network_isolated,
            "supported_formats": ["xlsx", "ods", "csv"],
        }

    def dry_run(
        self,
        *,
        snapshot: Mapping[str, Any],
        actions: tuple[Mapping[str, Any], ...],
        source_artifact: Mapping[str, Any] | None = None,
    ) -> Mapping[str, Any]:
        parsed = WorkbookSnapshotV1.from_mapping(snapshot)
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as exc:  # pragma: no cover - image/dependency gate
            raise SpreadsheetExecutionError("spreadsheet_openpyxl_unavailable") from exc

        with tempfile.TemporaryDirectory(prefix="ananta-spreadsheet-") as temporary:
            root = Path(temporary)
            source_dir = root / "source"
            result_dir = root / "result"
            import_dir = root / "imported"
            import_profile = root / "import-profile"
            profile = root / "recalc-profile"
            source_dir.mkdir(mode=0o700)
            result_dir.mkdir(mode=0o700)
            import_dir.mkdir(mode=0o700)
            import_profile.mkdir(mode=0o700)
            profile.mkdir(mode=0o700)
            source = source_dir / "workbook.xlsx"

            sheet_names: dict[str, str] = {}
            formula_asts: dict[tuple[str, str], Mapping[str, Any]] = {}
            style_refs: dict[tuple[str, str], str | None] = {}
            if source_artifact is None:
                workbook = Workbook()
                workbook.remove(workbook.active)
            else:
                if set(source_artifact) != {"content", "filename", "media_type", "sha256"}:
                    raise SpreadsheetExecutionError("spreadsheet_source_artifact_fields_invalid")
                content = source_artifact.get("content")
                if not isinstance(content, bytes):
                    raise SpreadsheetExecutionError("spreadsheet_source_artifact_content_invalid")
                inspection = self._inspector.inspect(
                    filename=str(source_artifact.get("filename") or ""),
                    media_type=str(source_artifact.get("media_type") or ""),
                    content=content,
                )
                if hashlib.sha256(content).hexdigest() != source_artifact.get("sha256"):
                    raise SpreadsheetExecutionError("spreadsheet_source_artifact_digest_invalid")
                imported_source = source_dir / f"input.{inspection.format}"
                imported_source.write_bytes(content)
                self._convert(source=imported_source, destination=import_dir, profile=import_profile)
                imported_workbook = import_dir / "input.xlsx"
                if not imported_workbook.is_file():
                    raise SpreadsheetExecutionError("spreadsheet_source_artifact_conversion_invalid")
                workbook = load_workbook(imported_workbook, data_only=False, read_only=False)
            for sheet in parsed.sheets:
                sheet_id = str(sheet["sheet_id"])
                name = str(sheet["name"])
                if source_artifact is None:
                    target = workbook.create_sheet(name)
                    target.sheet_state = "hidden" if sheet["hidden"] else "visible"
                elif name not in workbook.sheetnames:
                    raise SpreadsheetExecutionError("spreadsheet_source_snapshot_binding_invalid")
                sheet_names[sheet_id] = name
            if source_artifact is not None and set(workbook.sheetnames) != set(sheet_names.values()):
                raise SpreadsheetExecutionError("spreadsheet_source_snapshot_binding_invalid")
            for sheet in parsed.sheets:
                sheet_id = str(sheet["sheet_id"])
                target = workbook[sheet_names[sheet_id]]
                for cell in sheet["cells"]:
                    address = str(cell["address"])
                    if cell["formula"] is not None:
                        formula_asts[(sheet_id, address)] = copy.deepcopy(cell["formula"])
                        if source_artifact is None:
                            target[address] = "=" + render_formula(cell["formula"], sheet_names)
                    elif source_artifact is None:
                        target[address] = cell["value"]
                    style_refs[(sheet_id, address)] = cell["style_ref"]
            direct_targets = self._actions.apply(
                workbook=workbook,
                sheet_names=sheet_names,
                actions=actions,
                formula_asts=formula_asts,
                style_refs=style_refs,
            )
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
            workbook.save(source)
            workbook.close()

            self._convert(source=source, destination=result_dir, profile=profile)
            converted = result_dir / source.name
            if not converted.is_file() or converted.stat().st_size > self._file_bytes:
                raise SpreadsheetExecutionError("spreadsheet_libreoffice_output_invalid")
            formulas = load_workbook(converted, data_only=False, read_only=True)
            values = load_workbook(converted, data_only=True, read_only=True)
            try:
                candidate = _snapshot_from_workbooks(
                    parsed=parsed,
                    formulas=formulas,
                    values=values,
                    sheet_names=sheet_names,
                    formula_asts=formula_asts,
                    style_refs=style_refs,
                    actions=actions,
                )
            finally:
                formulas.close()
                values.close()
            result_content = converted.read_bytes()
            if len(result_content) > 16 * 1024 * 1024:
                raise SpreadsheetExecutionError("spreadsheet_result_artifact_too_large")

        normalized = WorkbookSnapshotV1.from_mapping(candidate)
        before = _cells(parsed.to_dict())
        after = _cells(normalized.to_dict())
        diffs = []
        for key in sorted(set(before) | set(after)):
            if before.get(key) == after.get(key):
                continue
            diffs.append(
                {
                    "action_id": direct_targets[key][-1] if key in direct_targets else None,
                    "action_ids": list(direct_targets.get(key) or []),
                    "sheet_id": key[0],
                    "cell": key[1],
                    "before": before.get(key),
                    "after": after.get(key),
                    "direct": key in direct_targets,
                }
            )
        return {
            "schema": "ananta.spreadsheet-execution-result.v1",
            "candidate_snapshot": normalized.to_dict(),
            "candidate_snapshot_digest": normalized.digest,
            "diff": diffs,
            "recalculation_performed": True,
            "engine": "libreoffice-calc",
            "engine_version": self._version,
            "production_fidelity": self._network_isolated,
            "result_artifact": {
                "content_base64": base64.b64encode(result_content).decode("ascii"),
                "sha256": hashlib.sha256(result_content).hexdigest(),
                "size_bytes": len(result_content),
                "format": "xlsx",
                "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            "human_intervention_required": False,
        }

    def import_document(
        self,
        *,
        content: bytes,
        filename: str,
        media_type: str,
        document_version_id: str,
    ) -> Mapping[str, Any]:
        inspection = self._inspector.inspect(filename=filename, media_type=media_type, content=content)
        digest = hashlib.sha256(content).hexdigest()
        with tempfile.TemporaryDirectory(prefix="ananta-spreadsheet-import-") as temporary:
            root = Path(temporary)
            source_dir = root / "source"
            result_dir = root / "result"
            profile = root / "profile"
            source_dir.mkdir(mode=0o700)
            result_dir.mkdir(mode=0o700)
            profile.mkdir(mode=0o700)
            source = source_dir / f"workbook.{inspection.format}"
            source.write_bytes(content)
            self._convert(source=source, destination=result_dir, profile=profile)
            converted = result_dir / "workbook.xlsx"
            if not converted.is_file() or converted.stat().st_size > self._file_bytes:
                raise SpreadsheetExecutionError("spreadsheet_libreoffice_output_invalid")
            snapshot, unsupported = _snapshot_from_imported_workbook(
                converted,
                source_digest=digest,
                document_version_id=document_version_id,
            )
        parsed = WorkbookSnapshotV1.from_mapping(snapshot)
        unsupported_objects = sorted({*inspection.unsupported_parts, *unsupported})
        return {
            "schema": "ananta.spreadsheet-import-result.v1",
            "snapshot": parsed.to_dict(),
            "snapshot_digest": parsed.digest,
            "source": {
                "sha256": digest,
                "size_bytes": len(content),
                "format": inspection.format,
                "media_type": inspection.media_type,
                "archive_entries": inspection.archive_entries,
                "uncompressed_bytes": inspection.uncompressed_bytes,
            },
            "unsupported_objects": unsupported_objects,
            "engine": "libreoffice-calc",
            "engine_version": self._version,
            "production_fidelity": self._network_isolated,
            "human_intervention_required": False,
        }

    def _convert(self, *, source: Path, destination: Path, profile: Path) -> None:
        command = [
            self._executable,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            "--invisible",
            f"-env:UserInstallation={profile.as_uri()}",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(destination),
            str(source),
        ]
        try:
            completed = subprocess.run(  # noqa: S603 - fixed executable and argv, no shell
                command,
                check=False,
                stdin=subprocess.DEVNULL,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                timeout=self._timeout,
                env={
                    "PATH": "/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin",
                    "LANG": "C.UTF-8",
                    "LC_ALL": "C.UTF-8",
                    "TZ": "UTC",
                },
                start_new_session=True,
                preexec_fn=self._limits,
            )
        except subprocess.TimeoutExpired as exc:
            raise SpreadsheetExecutionError("spreadsheet_libreoffice_timeout") from exc
        if completed.returncode != 0:
            raise SpreadsheetExecutionError("spreadsheet_libreoffice_failed")

    def _limits(self) -> None:
        resource.setrlimit(resource.RLIMIT_FSIZE, (self._file_bytes, self._file_bytes))
        resource.setrlimit(resource.RLIMIT_AS, (self._memory_bytes, self._memory_bytes))
        resource.setrlimit(resource.RLIMIT_NOFILE, (128, 128))

    def _detect_version(self) -> str:
        try:
            completed = subprocess.run(  # noqa: S603 - fixed executable and argv, no shell
                [self._executable, "--headless", "--version"],
                check=False,
                capture_output=True,
                text=True,
                timeout=10,
            )
        except (OSError, subprocess.SubprocessError) as exc:
            raise SpreadsheetExecutionError("spreadsheet_libreoffice_unavailable") from exc
        value = " ".join(completed.stdout.split())
        if completed.returncode != 0 or not value or len(value) > 128:
            raise SpreadsheetExecutionError("spreadsheet_libreoffice_version_invalid")
        return value


def _snapshot_from_workbooks(
    *,
    parsed: WorkbookSnapshotV1,
    formulas: Any,
    values: Any,
    sheet_names: Mapping[str, str],
    formula_asts: Mapping[tuple[str, str], Mapping[str, Any]],
    style_refs: Mapping[tuple[str, str], str | None],
    actions: tuple[Mapping[str, Any], ...],
) -> dict[str, Any]:
    from openpyxl.utils import get_column_letter

    sheets = []
    for original in parsed.sheets:
        sheet_id = str(original["sheet_id"])
        formula_sheet = formulas[sheet_names[sheet_id]]
        value_sheet = values[sheet_names[sheet_id]]
        cells = []
        for row_index, row in enumerate(formula_sheet.iter_rows(), start=1):
            for column_index, raw_cell in enumerate(row, start=1):
                address = getattr(raw_cell, "coordinate", f"{get_column_letter(column_index)}{row_index}")
                computed = value_sheet[address].value
                formula_ast = formula_asts.get((sheet_id, address))
                if raw_cell.value is None and computed is None:
                    continue
                cells.append(
                    {
                        "address": address,
                        "value": computed if formula_ast is not None else raw_cell.value,
                        "formula": copy.deepcopy(formula_ast),
                        "style_ref": style_refs.get((sheet_id, address)),
                    }
                )
        sheets.append(
            {
                "sheet_id": sheet_id,
                "name": str(original["name"]),
                "hidden": bool(original["hidden"]),
                "cells": cells,
            }
        )
    return {
        "schema": WorkbookSnapshotV1.SCHEMA,
        "snapshot_id": f"libreoffice-{canonical_digest(list(actions))[:24]}",
        "document_version_id": parsed.document_version_id,
        "sheets": sheets,
    }


def _cells(snapshot: Mapping[str, Any]) -> dict[tuple[str, str], Mapping[str, Any]]:
    return {
        (str(sheet["sheet_id"]), str(cell["address"])): cell for sheet in snapshot["sheets"] for cell in sheet["cells"]
    }


def _snapshot_from_imported_workbook(
    path: Path,
    *,
    source_digest: str,
    document_version_id: str,
) -> tuple[dict[str, Any], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - image/dependency gate
        raise SpreadsheetExecutionError("spreadsheet_openpyxl_unavailable") from exc
    formulas = load_workbook(path, data_only=False, read_only=False)
    values = load_workbook(path, data_only=True, read_only=False)
    try:
        if not 1 <= len(formulas.worksheets) <= 64:
            raise SpreadsheetExecutionError("spreadsheet_sheet_count_invalid")
        sheet_ids = {
            sheet.title.casefold(): f"sheet-{source_digest[:16]}-{index + 1}"
            for index, sheet in enumerate(formulas.worksheets)
        }
        sheets = []
        unsupported: list[str] = []
        cell_count = 0
        for formula_sheet, value_sheet in zip(formulas.worksheets, values.worksheets, strict=True):
            sheet_id = sheet_ids[formula_sheet.title.casefold()]
            if formula_sheet.max_row * formula_sheet.max_column > 100_000:
                raise SpreadsheetExecutionError("spreadsheet_cell_limit_exceeded")
            cells = []
            for row in formula_sheet.iter_rows():
                for formula_cell in row:
                    computed = value_sheet[formula_cell.coordinate].value
                    if formula_cell.value is None and computed is None:
                        continue
                    cell_count += 1
                    if cell_count > 100_000:
                        raise SpreadsheetExecutionError("spreadsheet_cell_limit_exceeded")
                    formula_ast = None
                    cell_value = _json_cell(computed if formula_cell.data_type == "f" else formula_cell.value)
                    if formula_cell.data_type == "f":
                        try:
                            formula_ast = parse_formula(
                                str(formula_cell.value),
                                current_sheet_id=sheet_id,
                                sheet_ids_by_name=sheet_ids,
                            )
                        except SpreadsheetFormulaUnsupported:
                            unsupported.append(f"formula:{sheet_id}:{formula_cell.coordinate}")
                    cells.append(
                        {
                            "address": formula_cell.coordinate,
                            "value": cell_value,
                            "formula": formula_ast,
                            "style_ref": f"style-{formula_cell.style_id}" if formula_cell.style_id else None,
                        }
                    )
            sheets.append(
                {
                    "sheet_id": sheet_id,
                    "name": formula_sheet.title,
                    "hidden": formula_sheet.sheet_state != "visible",
                    "cells": cells,
                }
            )
        return (
            {
                "schema": WorkbookSnapshotV1.SCHEMA,
                "snapshot_id": f"import-{source_digest[:24]}",
                "document_version_id": document_version_id,
                "sheets": sheets,
            },
            unsupported,
        )
    finally:
        formulas.close()
        values.close()


def _json_cell(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value)


__all__ = ["LibreOfficeSpreadsheetExecutor", "SpreadsheetExecutionError"]
