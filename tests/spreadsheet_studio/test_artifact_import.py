from __future__ import annotations

import hashlib
import io
import shutil
import subprocess
import zipfile
from pathlib import Path

import pytest

from agent.services.spreadsheet_artifact_store import SpreadsheetArtifactStore
from agent.services.spreadsheet_policy import SpreadsheetPolicy
from agent.services.spreadsheet_saga_service import SpreadsheetSagaService
from agent.services.spreadsheet_store import SpreadsheetStore
from worker.spreadsheet.artifact_inspector import SpreadsheetArtifactInspector, SpreadsheetArtifactRejected
from worker.spreadsheet.formula_parser import SpreadsheetFormulaUnsupported, parse_formula
from worker.spreadsheet.libreoffice_executor import LibreOfficeSpreadsheetExecutor


def _xlsx_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Budget"
    sheet["A1"] = 41
    sheet["B1"] = "=A1+1"
    sheet["C1"] = "preserved-style"
    sheet["C1"].font = Font(bold=True)
    sheet["C1"].fill = PatternFill(fill_type="solid", fgColor="FFCC00")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_artifact_inspector_rejects_spoofing_traversal_active_content_and_csv_injection() -> None:
    inspector = SpreadsheetArtifactInspector()
    xlsx = _xlsx_bytes()
    with pytest.raises(SpreadsheetArtifactRejected, match="media_type_mismatch"):
        inspector.inspect(filename="book.xlsx", media_type="text/csv", content=xlsx)

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w") as value:
        value.writestr("[Content_Types].xml", "safe")
        value.writestr("../escape.xml", "unsafe")
    with pytest.raises(SpreadsheetArtifactRejected, match="archive_path_invalid"):
        inspector.inspect(
            filename="book.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=archive.getvalue(),
        )

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w") as value:
        value.writestr("[Content_Types].xml", "safe")
        value.writestr("xl/vbaProject.bin", "unsafe")
    with pytest.raises(SpreadsheetArtifactRejected, match="active_content_forbidden"):
        inspector.inspect(
            filename="book.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=archive.getvalue(),
        )

    with pytest.raises(SpreadsheetArtifactRejected, match="csv_formula_injection_forbidden"):
        inspector.inspect(filename="book.csv", media_type="text/csv", content=b"name,value\nattack,=CMD()\n")
    accepted = inspector.inspect(filename="book.csv", media_type="text/csv", content=b"name,value\ndelta,-42.5\n")
    assert accepted.format == "csv"


def test_immutable_artifact_store_uses_opaque_identity_and_verifies_reads(tmp_path: Path) -> None:
    content = _xlsx_bytes()
    digest = hashlib.sha256(content).hexdigest()
    store = SpreadsheetArtifactStore(tmp_path)
    first = store.store(
        tenant_id="tenant-a",
        content=content,
        format="xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        expected_sha256=digest,
    )
    replay = store.store(
        tenant_id="tenant-a",
        content=content,
        format="xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        expected_sha256=digest,
    )
    assert first == replay
    assert first.artifact_id == f"artifact-{digest[:32]}"
    assert store.read(tenant_id="tenant-a", sha256=digest, format="xlsx") == content
    with pytest.raises(KeyError, match="not_found"):
        store.read(tenant_id="tenant-b", sha256=digest, format="xlsx")


def test_formula_parser_is_closed_and_sheet_bound() -> None:
    mapping = {"budget": "sheet-budget"}
    assert parse_formula("=A1+1", current_sheet_id="sheet-budget", sheet_ids_by_name=mapping) == {
        "op": "add",
        "left": {"op": "cell", "sheet_id": "sheet-budget", "cell": "A1"},
        "right": {"op": "literal", "value": 1},
    }
    assert (
        parse_formula("=SUM('Budget'!A1:A4)", current_sheet_id="sheet-budget", sheet_ids_by_name=mapping)["op"]
        == "sum_range"
    )
    with pytest.raises(SpreadsheetFormulaUnsupported, match="unsupported"):
        parse_formula('=WEBSERVICE("https://example.test")', current_sheet_id="sheet-budget", sheet_ids_by_name=mapping)


@pytest.mark.integration
@pytest.mark.parametrize(
    ("format", "media_type"),
    [
        ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("ods", "application/vnd.oasis.opendocument.spreadsheet"),
        ("csv", "text/csv"),
    ],
)
def test_real_libreoffice_imports_supported_formats(tmp_path: Path, format: str, media_type: str) -> None:
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if not executable:
        pytest.skip("LibreOffice is not installed")
    if format == "csv":
        content = b"value,label\n41,safe\n"
    elif format == "xlsx":
        content = _xlsx_bytes()
    else:
        source = tmp_path / "source.xlsx"
        source.write_bytes(_xlsx_bytes())
        completed = subprocess.run(  # noqa: S603 - test invokes the resolved fixed LibreOffice executable
            [executable, "--headless", "--convert-to", "ods", "--outdir", str(tmp_path), str(source)],
            check=False,
            capture_output=True,
            timeout=30,
        )
        assert completed.returncode == 0
        content = (tmp_path / "source.ods").read_bytes()
    result = LibreOfficeSpreadsheetExecutor(network_isolated=True).import_document(
        content=content,
        filename=f"workbook.{format}",
        media_type=media_type,
        document_version_id="version-one",
    )
    assert result["source"]["format"] == format
    assert result["snapshot_digest"]
    assert result["human_intervention_required"] is False
    assert result["production_fidelity"] is True
    assert result["snapshot"]["sheets"]


@pytest.mark.integration
def test_imported_original_is_immutable_and_promoted_xlsx_preserves_unmodified_style(tmp_path: Path) -> None:
    if not (shutil.which("libreoffice") or shutil.which("soffice")):
        pytest.skip("LibreOffice is not installed")
    artifacts = SpreadsheetArtifactStore(tmp_path / "artifacts")
    studio = SpreadsheetSagaService(
        SpreadsheetStore(tmp_path / "documents.sqlite3"),
        policy=SpreadsheetPolicy(enabled=True, mode="worker", automatic_promotion_enabled=True),
        executor=LibreOfficeSpreadsheetExecutor(network_isolated=True),
        artifact_store=artifacts,
    )
    original = _xlsx_bytes()
    document = studio.import_document(
        tenant_id="tenant-a",
        owner_id="user-a",
        title="Budget",
        filename="budget.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=original,
        document_id="document-one",
    )
    sheet_id = document["snapshot"]["sheets"][0]["sheet_id"]
    result = studio.execute_proposal(
        tenant_id="tenant-a",
        principal_id="user-a",
        proposal={
            "schema": "ananta.spreadsheet-proposal.v1",
            "proposal_id": "proposal-one",
            "document_id": document["document_id"],
            "expected_version": document["version"],
            "base_snapshot_digest": document["snapshot_digest"],
            "actions": [
                {
                    "action_id": "set-value",
                    "kind": "set_value",
                    "sheet_id": sheet_id,
                    "cell": "A1",
                    "value": 42,
                    "formula": None,
                }
            ],
            "validators": [
                {
                    "validator_id": "equals",
                    "kind": "equals",
                    "sheet_id": sheet_id,
                    "cell": "A1",
                    "expected": 42,
                    "minimum": None,
                    "maximum": None,
                }
            ],
            "automatic_promotion": True,
        },
    )

    published, artifact = studio.download_published(
        tenant_id="tenant-a", document_id="document-one", principal_id="user-a"
    )
    unchanged_original, _ = studio.download_original(
        tenant_id="tenant-a", document_id="document-one", principal_id="user-a"
    )
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(published), data_only=False)
    try:
        assert workbook["Budget"]["A1"].value == 42
        assert workbook["Budget"]["C1"].font.bold is True
        assert workbook["Budget"]["C1"].fill.fgColor.rgb.endswith("FFCC00")
    finally:
        workbook.close()
    assert result["state"] == "promoted"
    assert result["candidate_artifact"]["sha256"] == artifact["sha256"]
    assert unchanged_original == original
    assert hashlib.sha256(published).hexdigest() != hashlib.sha256(original).hexdigest()
