"""Bounded CSV/XLSX extraction that never evaluates workbook content."""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from pathlib import PurePosixPath
from typing import Protocol

from openpyxl import load_workbook

from agent.services.business_controlling_import_service import (
    BusinessControllingImportError,
    TabularProfileRequest,
    WorkbookRiskMetadata,
)


class ControllingArtifactPort(Protocol):
    def read_bytes(self, *, tenant_id: str, project_id: str, source_revision_id: str) -> bytes: ...


@dataclass(frozen=True)
class TabularExtractionRequest:
    tenant_id: str
    project_id: str
    source_revision_id: str
    revision_digest: str
    source_format: str
    sheet_name: str | None = None


class BusinessControllingTabularExtractor:
    MAX_SOURCE_BYTES = 16 * 1024 * 1024
    MAX_ARCHIVE_ENTRIES = 2_000
    MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
    MAX_EXPANSION_RATIO = 100

    def __init__(self, artifacts: ControllingArtifactPort) -> None:
        self._artifacts = artifacts

    def extract(self, request: TabularExtractionRequest) -> TabularProfileRequest:
        payload = self._artifacts.read_bytes(
            tenant_id=request.tenant_id,
            project_id=request.project_id,
            source_revision_id=request.source_revision_id,
        )
        if not isinstance(payload, bytes) or not payload or len(payload) > self.MAX_SOURCE_BYTES:
            raise BusinessControllingImportError("controlling_source_size_invalid")
        if request.source_format == "csv":
            headers, rows, risk = self._csv(payload)
        elif request.source_format == "xlsx":
            headers, rows, risk = self._xlsx(payload, sheet_name=request.sheet_name)
        else:
            raise BusinessControllingImportError("controlling_source_format_denied")
        return TabularProfileRequest(
            tenant_id=request.tenant_id,
            project_id=request.project_id,
            source_revision_id=request.source_revision_id,
            revision_digest=request.revision_digest,
            source_format=request.source_format,
            headers=headers,
            rows=rows,
            risk=risk,
        )

    @staticmethod
    def _csv(payload: bytes) -> tuple[tuple[str, ...], tuple[tuple[object, ...], ...], WorkbookRiskMetadata]:
        try:
            text = payload.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise BusinessControllingImportError("controlling_csv_encoding_invalid") from exc
        if "\x00" in text:
            raise BusinessControllingImportError("controlling_csv_content_invalid")
        parsed = tuple(tuple(cell for cell in row) for row in csv.reader(io.StringIO(text, newline="")))
        if not parsed:
            raise BusinessControllingImportError("controlling_headers_invalid")
        headers = tuple(parsed[0])
        rows = tuple(parsed[1:])
        formula_cells = any(_looks_executable_csv_cell(cell) for row in rows for cell in row)
        return headers, rows, WorkbookRiskMetadata(has_formula_cells=formula_cells)

    def _xlsx(
        self, payload: bytes, *, sheet_name: str | None
    ) -> tuple[tuple[str, ...], tuple[tuple[object, ...], ...], WorkbookRiskMetadata]:
        risk = self._xlsx_preflight(payload)
        try:
            workbook = load_workbook(
                io.BytesIO(payload),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
            raise BusinessControllingImportError("controlling_xlsx_invalid") from exc
        try:
            visible = tuple(sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible")
            if sheet_name is None:
                if len(visible) != 1:
                    raise BusinessControllingImportError("controlling_xlsx_sheet_selection_required")
                sheet = visible[0]
            else:
                if sheet_name not in workbook.sheetnames:
                    raise BusinessControllingImportError("controlling_xlsx_sheet_unknown")
                sheet = workbook[sheet_name]
                if sheet.sheet_state != "visible":
                    raise BusinessControllingImportError("controlling_xlsx_hidden_sheet_denied")
            extracted: list[tuple[object, ...]] = []
            formula_cells = risk.has_formula_cells
            for row in sheet.iter_rows():
                formula_cells = formula_cells or any(cell.data_type == "f" for cell in row)
                extracted.append(tuple(cell.value for cell in row))
            if not extracted:
                raise BusinessControllingImportError("controlling_headers_invalid")
            return (
                tuple(str(value or "") for value in extracted[0]),
                tuple(extracted[1:]),
                WorkbookRiskMetadata(
                    has_macros=risk.has_macros,
                    has_external_links=risk.has_external_links,
                    has_formula_cells=formula_cells,
                    has_unsupported_objects=risk.has_unsupported_objects,
                ),
            )
        finally:
            workbook.close()

    def _xlsx_preflight(self, payload: bytes) -> WorkbookRiskMetadata:
        try:
            with zipfile.ZipFile(io.BytesIO(payload)) as archive:
                entries = archive.infolist()
                if len(entries) > self.MAX_ARCHIVE_ENTRIES:
                    raise BusinessControllingImportError("controlling_xlsx_archive_budget_exceeded")
                total = sum(entry.file_size for entry in entries)
                compressed = max(1, sum(entry.compress_size for entry in entries))
                if total > self.MAX_UNCOMPRESSED_BYTES or total / compressed > self.MAX_EXPANSION_RATIO:
                    raise BusinessControllingImportError("controlling_xlsx_archive_budget_exceeded")
                names = tuple(entry.filename.casefold() for entry in entries)
                if any(_unsafe_archive_name(entry.filename) for entry in entries):
                    raise BusinessControllingImportError("controlling_xlsx_path_traversal_denied")
        except zipfile.BadZipFile as exc:
            raise BusinessControllingImportError("controlling_xlsx_invalid") from exc
        return WorkbookRiskMetadata(
            has_macros=any(name.endswith("vbaproject.bin") for name in names),
            has_external_links=any("/externallinks/" in name for name in names),
            has_unsupported_objects=any(
                marker in name
                for name in names
                for marker in ("/embeddings/", "/activex/", "/oleobjects/", "/connections.xml")
            ),
        )


def _unsafe_archive_name(name: str) -> bool:
    path = PurePosixPath(name.replace("\\", "/"))
    return path.is_absolute() or ".." in path.parts


def _looks_executable_csv_cell(value: str) -> bool:
    return value.lstrip().startswith(("=", "+", "@", "\t", "\r"))


__all__ = [
    "BusinessControllingTabularExtractor",
    "ControllingArtifactPort",
    "TabularExtractionRequest",
]
